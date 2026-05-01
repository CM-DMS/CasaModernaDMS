"""Import CM Products from an Excel file.

Each row must have at least: item_name, item_group, cm_supplier_name.
Optional: name (existing item code — skipped if already present),
          cm_supplier_code, cm_rrp_ex_vat, cm_offer_tier1_inc_vat,
          cm_purchase_price_ex_vat, cm_shipping_fee, cm_handling_fee,
          cm_vat_rate_percent.

Run via bench:
    cd /home/frappe/frappe/casamoderna-bench-v3
    bench --site cms.local execute \
        casamoderna_dms.tools.import_products_xlsx.run \
        --kwargs '{"xlsx_path": "/home/frappe/products_2026-05-01.xlsx"}'

Add dry_run=true to preview without writing:
    --kwargs '{"xlsx_path": "...", "dry_run": true}'
"""
from __future__ import annotations

import frappe


def run(xlsx_path: str, dry_run: bool = False) -> None:
    try:
        import openpyxl
    except ImportError:
        frappe.throw("openpyxl is required: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    created = 0
    skipped = 0
    errors: list[str] = []

    for row_idx in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)}

        item_name = (row.get("item_name") or "").strip()
        if not item_name:
            continue

        existing_code = (row.get("name") or "").strip()
        if existing_code and frappe.db.exists("CM Product", existing_code):
            print(f"  SKIP  {existing_code} — already exists ({item_name})")
            skipped += 1
            continue

        # Derive supplier code if not provided
        supplier_code = (row.get("cm_supplier_code") or "").strip().upper()
        if not supplier_code:
            # Infer from supplier name map or fall back to first 3 letters
            supplier_name = (row.get("cm_supplier_name") or "").strip()
            supplier_code = _infer_supplier_code(supplier_name)

        doc_data = {
            "doctype":              "CM Product",
            "item_name":            item_name,
            "cm_given_name":        (row.get("cm_given_name") or item_name).strip(),
            "item_group":           (row.get("item_group") or "Tiles").strip(),
            "cm_product_type":      (row.get("cm_product_type") or "Primary").strip(),
            "cm_supplier_name":     (row.get("cm_supplier_name") or "").strip(),
            "cm_supplier_code":     supplier_code,
            "cm_vat_rate_percent":  float(row.get("cm_vat_rate_percent") or 18),
            # Cost inputs
            "cm_purchase_price_ex_vat": float(row.get("cm_purchase_price_ex_vat") or 0),
            "cm_shipping_fee":          float(row.get("cm_shipping_fee") or 0),
            "cm_handling_fee":          float(row.get("cm_handling_fee") or 0),
            "cm_shipping_percent":      float(row.get("cm_shipping_percent") or 0),
            "cm_other_landed":          float(row.get("cm_other_landed") or 0),
            # RRP: set cm_rrp_ex_vat directly; cm_target_margin_percent=0 means
            # _compute_pricing won't overwrite cm_rrp_ex_vat
            "cm_target_margin_percent": float(row.get("cm_target_margin_percent") or 0),
            "cm_rrp_ex_vat":            _rrp_ex_from_row(row),
            # Offer tier 1 (set directly; compute_pricing will fill ex_vat + disc%)
            "cm_offer_tier1_inc_vat": float(row.get("cm_offer_tier1_inc_vat") or 0),
        }

        if dry_run:
            print(f"  DRY   would create: {item_name!r} (supplier_code={supplier_code})")
            created += 1
            continue

        try:
            doc = frappe.get_doc(doc_data)
            doc.insert(ignore_permissions=True)
            frappe.db.commit()
            print(f"  OK    {doc.name} — {item_name}")
            created += 1
        except Exception as exc:
            msg = f"{item_name}: {exc}"
            print(f"  ERR   {msg}")
            errors.append(msg)
            frappe.db.rollback()

    wb.close()

    print()
    print(f"Done.  created={created}  skipped={skipped}  errors={len(errors)}")
    if errors:
        print("Errors:")
        for e in errors:
            print(f"  {e}")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_SUPPLIER_CODE_MAP: dict[str, str] = {
    "sotemail":  "SOT",
    "arezia":    "ARE",
    "milorex":   "MIL",
}


def _infer_supplier_code(supplier_name: str) -> str:
    """Return a 3-letter supplier code from the supplier name."""
    lower = supplier_name.lower()
    for key, code in _SUPPLIER_CODE_MAP.items():
        if key in lower:
            return code
    # Fall back: first 3 alpha chars
    alpha = "".join(c for c in supplier_name.upper() if c.isalpha())
    return alpha[:3] if len(alpha) >= 3 else "UNK"


def _rrp_ex_from_row(row: dict) -> float:
    """Return cm_rrp_ex_vat from the row.

    Prefer an explicit cm_rrp_ex_vat column; otherwise back-calculate
    from cm_rrp_inc_vat using the row's VAT rate.
    """
    if row.get("cm_rrp_ex_vat"):
        return float(row["cm_rrp_ex_vat"])
    rrp_inc = float(row.get("cm_rrp_inc_vat") or 0)
    if rrp_inc:
        vat = float(row.get("cm_vat_rate_percent") or 18)
        return round(rrp_inc / (1 + vat / 100), 9)
    return 0.0
